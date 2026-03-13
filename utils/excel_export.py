"""Excel 산출물 생성 유틸리티"""
import io
from typing import Dict, List, Any

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


# 스타일 상수
COLOR_HEADER_BG = "2F5496"   # 진한 파랑 (헤더)
COLOR_SECTION_BG = "D9E1F2"  # 연한 파랑 (섹션)
COLOR_ODD_COL = "E2EFDA"     # 연한 초록 (Core ODD 열)
COLOR_EXT_COL = "FFF2CC"     # 연한 노랑 (Extension 열)
COLOR_QTY_COL = "FCE4D6"     # 연한 주황 (Qty/Priority 열)

FONT_WHITE = Font(color="FFFFFF", bold=True)
FONT_BOLD = Font(bold=True)


def _header_style(ws, row, col, value, bg_color=COLOR_HEADER_BG, font=None):
    cell = ws.cell(row=row, column=col, value=value)
    cell.fill = PatternFill(fill_type="solid", fgColor=bg_color)
    cell.font = font or FONT_WHITE
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    return cell


def _set_border(ws, min_row, max_row, min_col, max_col):
    thin = Side(style="thin")
    for row in ws.iter_rows(min_row=min_row, max_row=max_row, min_col=min_col, max_col=max_col):
        for cell in row:
            cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)


def _build_value_label_map(catalog: Dict) -> Dict[str, Dict[str, str]]:
    """attribute_key → {value_code: label_kor} 매핑 생성"""
    mapping = {}
    for super_data in catalog.get("super_classes", {}).values():
        for class_data in super_data.get("classes", {}).values():
            for attr_key, attr_data in class_data.get("attributes", {}).items():
                vmap = {}
                for v in attr_data.get("values", []):
                    vmap[v["value_code"]] = v.get("label_kor") or v["value_code"]
                mapping[attr_key] = vmap
    return mapping


def generate_excel(output_data: Dict[str, Any], catalog: Dict) -> io.BytesIO:
    """
    요구사항 데이터를 Excel로 변환하여 BytesIO 반환

    Args:
        output_data: submit 시 생성된 JSON 데이터
        catalog: ODD catalog (value 한글 라벨 변환용)

    Returns:
        Excel 파일이 담긴 BytesIO 객체
    """
    wb = openpyxl.Workbook()
    value_label_map = _build_value_label_map(catalog)

    _write_summary_sheet(wb, output_data)
    _write_scenario_sheet(wb, output_data, value_label_map)
    _write_selection_sheet(wb, output_data, value_label_map)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


def _write_summary_sheet(wb: openpyxl.Workbook, data: Dict):
    """Sheet 1: 요청 요약"""
    ws = wb.active
    ws.title = "요청 요약"

    req = data["request"]
    summary = data.get("summary", {})

    qty_unit = summary.get("qty_unit", "frame")

    # feature 정보 수집 (per-feature 합산용)
    feature_names = []
    for g in data.get("scenario_groups", []):
        for combo in g.get("combinations", []):
            for fn in combo.get("features", {}).keys():
                if fn not in feature_names:
                    feature_names.append(fn)
    use_features = (qty_unit == "frame") and bool(feature_names)

    requester = req.get("requester", {})
    rows = [
        ("항목", "내용"),
        ("Request ID", req.get("id", "")),
        ("제목", req.get("title", "")),
        ("설명", req.get("description", "")),
        ("요청인 이름", requester.get("name", "")),
        ("요청인 이메일", requester.get("email", "")),
        ("Product", req.get("product", "")),
        ("Scenario 타입", req.get("scenario", "")),
        ("요청 범위", ", ".join(req.get("scope", []))),
        ("희망 완료일", req.get("due_date") or "미지정"),
        ("ODD Catalog 버전", req.get("catalog_version", "")),
        ("생성 일시", req.get("created_at", "")),
        ("", ""),
        ("총 시나리오 그룹 수", summary.get("total_groups", 0)),
        ("총 조건조합 수", summary.get("total_combinations", 0)),
        ("수량 단위", qty_unit),
    ]

    if use_features:
        for feat in feature_names:
            feat_total = sum(
                combo.get("features", {}).get(feat, 0)
                for g in data.get("scenario_groups", [])
                for combo in g.get("combinations", [])
            )
            rows.append((f"Feature: {feat}", f"{feat_total:,} frame"))
    else:
        rows.append(("총 Qty 합계", f"{summary.get('total_qty', 0):,}"))

    for r_idx, (label, value) in enumerate(rows, start=1):
        if label == "항목":
            _header_style(ws, r_idx, 1, label)
            _header_style(ws, r_idx, 2, value)
        elif label == "":
            pass
        else:
            c1 = ws.cell(row=r_idx, column=1, value=label)
            c1.font = FONT_BOLD
            c1.fill = PatternFill(fill_type="solid", fgColor=COLOR_SECTION_BG)
            ws.cell(row=r_idx, column=2, value=value)

    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 40
    _set_border(ws, 1, len(rows), 1, 2)


def _write_scenario_sheet(wb: openpyxl.Workbook, data: Dict, value_label_map: Dict):
    """Sheet 2: 조건조합 목록 (핵심 산출물)"""
    ws = wb.create_sheet("조건조합 목록")
    groups = data.get("scenario_groups", [])

    # 모든 조건조합을 flat list로 (group_name, no, combo)
    all_combos = []
    for g in groups:
        for i, combo in enumerate(g.get("combinations", []), start=1):
            all_combos.append((g["name"], i, combo))

    if not all_combos:
        ws.cell(row=1, column=1, value="조건조합 없음")
        return

    # 모든 attribute key 수집
    all_attr_keys = []
    for _, _, combo in all_combos:
        for k in combo.get("attributes", {}).keys():
            if k not in all_attr_keys:
                all_attr_keys.append(k)

    region_keys = [k for k in all_attr_keys if k == "ext.region"]
    core_keys = [k for k in all_attr_keys if not k.startswith("ext.")]
    ext_keys = [k for k in all_attr_keys if k.startswith("ext.") and k != "ext.region" and k != "ext.region.detail"]
    ordered_keys = region_keys + core_keys + ext_keys

    qty_unit = data.get("summary", {}).get("qty_unit", "frame")

    # Feature 이름 수집
    feature_names = []
    for _, _, combo in all_combos:
        for fn in combo.get('features', {}).keys():
            if fn not in feature_names:
                feature_names.append(fn)

    # frame 단위 + feature 있으면 qty 열 제거, feature 열만 사용
    use_features = (qty_unit == "frame") and bool(feature_names)

    if use_features:
        headers = ["시나리오 그룹", "No", "scenario_id"] + ordered_keys + ["priority", "notes"] + feature_names
        col_types = (
            ["group", "no", "id"]
            + ["region"] * len(region_keys)
            + ["core"] * len(core_keys)
            + ["ext"] * len(ext_keys)
            + ["priority", "notes"]
            + ["feature"] * len(feature_names)
        )
    else:
        headers = ["시나리오 그룹", "No", "scenario_id"] + ordered_keys + [f"qty ({qty_unit})", "priority", "notes"]
        col_types = (
            ["group", "no", "id"]
            + ["region"] * len(region_keys)
            + ["core"] * len(core_keys)
            + ["ext"] * len(ext_keys)
            + ["qty", "priority", "notes"]
        )

    COLOR_MAP = {
        "group": "1F3864",
        "no": "2F5496",
        "id": "4472C4",
        "region": "5B9BD5",
        "core": COLOR_ODD_COL,
        "ext": COLOR_EXT_COL,
        "qty": COLOR_QTY_COL,
        "priority": COLOR_QTY_COL,
        "notes": "F2F2F2",
        "feature": "E8D5F5",
    }
    group_labels = {
        "group": "시나리오 그룹",
        "no": "No",
        "id": "ID",
        "region": "Region",
        "core": "Core ODD",
        "ext": "Extension",
        "qty": "수량",
        "priority": "우선순위",
        "notes": "메모",
        "feature": "Feature",
    }
    group_colors = {
        "group": "0D0D0D",
        "no": "1F3864",
        "id": "1F3864",
        "region": "2F5496",
        "core": "375623",
        "ext": "7F6000",
        "qty": "843C0C",
        "priority": "843C0C",
        "notes": "595959",
        "feature": "5C1B8A",
    }

    # Row 1: 그룹 헤더 (연속 동일 타입 merge)
    prev_type = None
    group_start = 1
    for col_idx, ctype in enumerate(col_types, start=1):
        if ctype != prev_type and prev_type is not None:
            label = group_labels.get(prev_type, prev_type)
            color = group_colors.get(prev_type, "595959")
            ws.merge_cells(start_row=1, start_column=group_start, end_row=1, end_column=col_idx - 1)
            cell = ws.cell(row=1, column=group_start, value=label)
            cell.fill = PatternFill(fill_type="solid", fgColor=color)
            cell.font = FONT_WHITE
            cell.alignment = Alignment(horizontal="center", vertical="center")
            group_start = col_idx
        prev_type = ctype
    if prev_type:
        label = group_labels.get(prev_type, prev_type)
        color = group_colors.get(prev_type, "595959")
        ws.merge_cells(start_row=1, start_column=group_start, end_row=1, end_column=len(col_types))
        cell = ws.cell(row=1, column=group_start, value=label)
        cell.fill = PatternFill(fill_type="solid", fgColor=color)
        cell.font = FONT_WHITE
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # Row 2: 상세 헤더
    for col_idx, (header, ctype) in enumerate(zip(headers, col_types), start=1):
        bg = COLOR_MAP.get(ctype, "CCCCCC")
        if ctype in ("group", "no", "id", "region"):
            _header_style(ws, 2, col_idx, header, bg_color=bg)
        else:
            cell = ws.cell(row=2, column=col_idx, value=header)
            cell.fill = PatternFill(fill_type="solid", fgColor=bg)
            cell.font = Font(bold=True, color="000000")
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # 데이터 행
    for row_idx, (group_name, no, combo) in enumerate(all_combos, start=3):
        attrs = combo.get("attributes", {})

        ws.cell(row=row_idx, column=1, value=group_name)
        ws.cell(row=row_idx, column=2, value=no)
        ws.cell(row=row_idx, column=3, value=combo.get("scenario_id", ""))

        for col_idx, key in enumerate(ordered_keys, start=4):
            raw_val = attrs.get(key, "")
            display_val = value_label_map.get(key, {}).get(raw_val, raw_val)
            if isinstance(display_val, str) and display_val.startswith("other:"):
                display_val = f"[기타] {display_val[6:]}"
            ws.cell(row=row_idx, column=col_idx, value=display_val)

        offset = len(ordered_keys) + 4
        if use_features:
            ws.cell(row=row_idx, column=offset, value=combo.get("priority", ""))
            ws.cell(row=row_idx, column=offset + 1, value=combo.get("notes", ""))
            for feat_idx, feat_name in enumerate(feature_names):
                feat_val = combo.get('features', {}).get(feat_name, 0)
                ws.cell(row=row_idx, column=offset + 2 + feat_idx, value=feat_val)
        else:
            ws.cell(row=row_idx, column=offset, value=combo.get("qty", 0))
            ws.cell(row=row_idx, column=offset + 1, value=combo.get("priority", ""))
            ws.cell(row=row_idx, column=offset + 2, value=combo.get("notes", ""))

        # 짝수 행 연한 배경
        if row_idx % 2 == 0:
            for c in range(1, len(headers) + 1):
                existing = ws.cell(row=row_idx, column=c).fill
                if existing.fill_type != "solid":
                    ws.cell(row=row_idx, column=c).fill = PatternFill(fill_type="solid", fgColor="F5F5F5")

    # 열 너비
    ws.column_dimensions["A"].width = 20   # 시나리오 그룹
    ws.column_dimensions["B"].width = 6    # No
    ws.column_dimensions["C"].width = 22   # scenario_id
    for col_idx in range(4, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = 18

    # 테두리 & 행 고정
    _set_border(ws, 1, 2 + len(all_combos), 1, len(headers))
    ws.freeze_panes = "D3"


def _write_selection_sheet(wb: openpyxl.Workbook, data: Dict, value_label_map: Dict):
    """Sheet 3: 선택 요약 (그룹별 Core ODD + Extension 선택 내용)"""
    ws = wb.create_sheet("선택 요약")
    groups = data.get("scenario_groups", [])
    row = 1

    for group_idx, g in enumerate(groups, start=1):
        sel = g.get("selection", {})

        # 그룹 헤더
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=3)
        cell = ws.cell(row=row, column=1, value=f"시나리오 그룹 {group_idx}: {g['name']}")
        cell.fill = PatternFill(fill_type="solid", fgColor="1F3864")
        cell.font = FONT_WHITE
        cell.alignment = Alignment(horizontal="center", vertical="center")
        row += 1

        # Core ODD 서브헤더
        _header_style(ws, row, 1, "Core ODD 선택 내용", bg_color=COLOR_HEADER_BG)
        _header_style(ws, row, 2, "선택 Value (코드)")
        _header_style(ws, row, 3, "선택 Value (한글)")
        row += 1

        for attr_key, value_codes in sel.get("core", {}).items():
            kor_labels = []
            for vc in value_codes:
                if vc.startswith("other:"):
                    kor_labels.append(f"[기타] {vc[6:]}")
                else:
                    kor_labels.append(value_label_map.get(attr_key, {}).get(vc, vc))
            c1 = ws.cell(row=row, column=1, value=attr_key)
            c1.font = FONT_BOLD
            ws.cell(row=row, column=2, value=", ".join(value_codes))
            ws.cell(row=row, column=3, value=", ".join(kor_labels))
            row += 1

        # Extension 서브헤더
        _header_style(ws, row, 1, "Extension 선택 내용", bg_color="7F6000")
        _header_style(ws, row, 2, "카테고리")
        _header_style(ws, row, 3, "값 목록")
        row += 1

        for ext_key, ext_data in sel.get("extensions", {}).items():
            c1 = ws.cell(row=row, column=1, value=ext_key)
            c1.font = FONT_BOLD
            ws.cell(row=row, column=2, value=ext_data.get("category", ""))
            if ext_key == "ext.region":
                region_val = ext_data.get("region_code", "")
                detail = ext_data.get("region_detail") or ""
                ws.cell(row=row, column=3, value=f"{region_val} {detail}".strip())
            else:
                ws.cell(row=row, column=3, value=", ".join(ext_data.get("values", [])))
            row += 1

        # 그룹 간 공백
        row += 1

    ws.column_dimensions["A"].width = 40
    ws.column_dimensions["B"].width = 30
    ws.column_dimensions["C"].width = 40
    _set_border(ws, 1, row - 1, 1, 3)
